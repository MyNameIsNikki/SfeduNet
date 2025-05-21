import sys
import yaml
import pandas as pd
import os
import logging
import time
import gc
import sqlalchemy as sa
from sqlalchemy import text  # Добавляем импорт для text()
import psycopg2
import mysql.connector
import traceback
import itertools
import xlsxwriter
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
from tenacity import retry, stop_after_attempt, wait_exponential
import pickle
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error
import psutil

# Логирование с поддержкой UTF-8
def get_logger():
    logger = logging.getLogger("ETL")
    logger.setLevel(logging.INFO)
    os.makedirs("logs", exist_ok=True)
    fh = logging.FileHandler("logs/etl.log", encoding='utf-8')
    formatter = logging.Formatter("%(asctime)s — %(levelname)s — %(message)s")
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    return logger

# Валидация миграции
def validate_migration(config, logger):
    source_type = config.get('source_type', 'excel')
    target_type = config.get('target_type', 'excel')
    
    try:
        if source_type == 'excel':
            src_df = pd.read_excel(config['source_file'], engine='openpyxl')
        else:
            engine = create_db_engine(source_type, config['db_params'])
            src_df = pd.read_sql(f"SELECT * FROM {config['db_params']['table']}", engine)

        if target_type == 'excel':
            tgt_df = pd.read_excel(config['target_file'], engine='openpyxl')
        else:
            engine = create_db_engine(target_type, config['db_params_target'])
            with engine.connect() as conn:
                table_exists = engine.dialect.has_table(conn, config['db_params_target']['table'])
            if not table_exists:
                logger.warning(f"⚠️ Target table '{config['db_params_target']['table']}' does not exist.")
                return False
            tgt_df = pd.read_sql(f"SELECT * FROM {config['db_params_target']['table']}", engine)

        src_count = len(src_df)
        tgt_count = len(tgt_df)
        logger.info(f"📊 Validation: Source rows = {src_count}, Target rows = {tgt_count}")
        
        if src_count != tgt_count:
            logger.warning(f"⚠️ Row count mismatch: Source={src_count}, Target={tgt_count}")
            return False
        
        src_df = src_df.sort_index(axis=1)
        tgt_df = tgt_df.sort_index(axis=1)
        if not src_df.equals(tgt_df):
            logger.warning("⚠️ Data integrity mismatch. Columns or values differ.")
            return False
        
        logger.info("✅ Validation passed: row count and data integrity.")
        return True
    except Exception as e:
        logger.error(f"❌ Validation failed: {e}")
        return False

# Создание SQLAlchemy движка с проверкой подключения
def create_db_engine(db_type, db_params):
    try:
        if db_type == 'postgresql':
            connection_string = (
                f"postgresql+psycopg2://{db_params['user']}:{db_params['password']}@"
                f"{db_params['host']}:{db_params['port']}/{db_params['database']}"
            )
        elif db_type == 'mysql':
            connection_string = (
                f"mysql+mysqlconnector://{db_params['user']}:{db_params['password']}@"
                f"{db_params['host']}:{db_params['port']}/{db_params['database']}"
            )
        else:
            raise ValueError(f"Unsupported database type: {db_type}")
        
        engine = sa.create_engine(connection_string)
        with engine.connect() as conn:
            # Используем text() для выполнения raw SQL
            conn.execute(text("SELECT 1"))
        return engine
    except sa.exc.OperationalError as e:
        raise Exception(
            f"Database connection failed: {str(e)}. "
            "Check host, port, database, user, password, and ensure the database server is running."
        )
    except Exception as e:
        raise Exception(f"Failed to create database engine: {str(e)}")

# Класс миграции с обучением
class ETLJob:
    def __init__(self, config):
        self.config = config
        self.source_type = config.get('source_type', 'excel')
        self.target_type = config.get('target_type', 'excel')
        self.source_file = config.get('source_file')
        self.target_file = config.get('target_file')
        self.db_params = config.get('db_params', {})
        self.db_params_target = config.get('db_params_target', {})
        self.chunk_size = min(config.get('chunk_size', 10000), 50000)  # Ограничение максимального чанка
        self.commit_interval = config.get('commit_interval', 20000)
        self.ga_pop_size = config.get('ga_pop_size', 20)
        self.ga_generations = config.get('ga_generations', 10)
        self.ga_weights = config.get('ga_weights', [10, 2, 2, 2, 0.5, 0.5])
        self.logger = get_logger()
        self.total_rows = 0
        self.total_rows_expected = 0
        self.progress_callback = config.get('progress_callback', None)
        self.duration = 0
        self.predicted_duration = 0

        # Инициализация модели и данных
        self.model_file = "etl_optimizer.pkl"
        self.data_file = "migration_logs.csv"
        self.model = self._load_or_init_model()
        self._ensure_data_file()

    def _load_or_init_model(self):
        if os.path.exists(self.model_file):
            with open(self.model_file, "rb") as f:
                return pickle.load(f)
        return RandomForestRegressor(n_estimators=100, random_state=42)

    def _ensure_data_file(self):
        if not os.path.exists(self.data_file):
            pd.DataFrame(columns=["file_size", "num_columns", "memory_available", "chunk_size", "commit_interval", "duration"]).to_csv(self.data_file, index=False)

    def _predict_parameters_and_duration(self):
        if self.source_type == 'excel':
            workbook = load_workbook(self.source_file, read_only=True)
            sheet = workbook.active
            num_rows = sheet.max_row - 1
            num_columns = len([cell.value for cell in next(sheet.rows) if cell.value])
            workbook.close()
        else:
            engine = create_db_engine(self.source_type, self.db_params)
            with engine.connect() as conn:
                result = conn.execute(text(f"SELECT COUNT(*) FROM {self.db_params['table']}"))
                num_rows = result.fetchone()[0]
                result = conn.execute(text(f"SELECT * FROM {self.db_params['table']} LIMIT 1"))
                num_columns = len(result.keys())

        memory_available = psutil.virtual_memory().available // (1024 * 1024)  # В МБ

        input_data = pd.DataFrame({
            "file_size": [num_rows],
            "num_columns": [num_columns],
            "memory_available": [memory_available]
        })

        try:
            predicted = self.model.predict(input_data)[0]
            self.chunk_size = max(100, int(predicted[0]))
            self.commit_interval = max(1000, int(predicted[1]))
            self.predicted_duration = max(1, int(predicted[2]))
        except Exception as e:
            self.logger.warning(f"Model not fitted, using default values: {e}")
            self.chunk_size = min(self.config.get('chunk_size', 10000), 50000)
            self.commit_interval = self.config.get('commit_interval', 20000)
            self.predicted_duration = max(1, int(num_rows / 1000))

        self.logger.info(f"Predicted parameters: chunk_size={self.chunk_size}, commit_interval={self.commit_interval}, duration={self.predicted_duration}s")

    def _log_migration_data(self):
        data = pd.read_csv(self.data_file)
        new_row = pd.DataFrame({
            "file_size": [self.total_rows_expected],
            "num_columns": [len(self.stream_data().__next__().columns) if self.total_rows_expected > 0 else 0],
            "memory_available": [psutil.virtual_memory().available // (1024 * 1024)],
            "chunk_size": [self.chunk_size],
            "commit_interval": [self.commit_interval],
            "duration": [self.duration]
        })
        data = pd.concat([data, new_row], ignore_index=True)
        data.to_csv(self.data_file, index=False)
        self.logger.info(f"Logged migration data.")

    def _train_model(self):
        data = pd.read_csv(self.data_file)
        if len(data) < 10:
            self.logger.warning("Not enough data for training. Need at least 10 records.")
            return

        X = data[["file_size", "num_columns", "memory_available"]]
        y = data[["chunk_size", "commit_interval", "duration"]]
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        self.model.fit(X_train, y_train)
        y_pred = self.model.predict(X_test)
        mse = mean_squared_error(y_test, y_pred)
        self.logger.info(f"Model retrained. Mean Squared Error: {mse}")

        with open(self.model_file, "wb") as f:
            pickle.dump(self.model, f)
        self.logger.info(f"Model saved.")

    def stream_data(self):
        if self.source_type == 'excel':
            if not os.path.exists(self.source_file):
                raise FileNotFoundError(f"Source file {self.source_file} does not exist.")
            if not os.access(self.source_file, os.R_OK):
                raise PermissionError(f"No read permission for {self.source_file}")
            workbook = load_workbook(self.source_file, read_only=True)
            sheet = workbook.active

            headers = [cell.value for cell in next(sheet.rows)]
            if not headers:
                raise ValueError("Excel file is empty or has no headers")

            self.total_rows_expected = sheet.max_row - 1
            self.logger.info(f"Total rows expected: {self.total_rows_expected}")

            chunk_data = []
            for row in sheet.rows:
                if row[0].row == 1:
                    continue
                row_data = [cell.value for cell in row]
                chunk_data.append(row_data)

                if len(chunk_data) >= self.chunk_size:
                    chunk = pd.DataFrame(chunk_data, columns=headers)
                    yield chunk
                    chunk_data = []

            if chunk_data:
                chunk = pd.DataFrame(chunk_data, columns=headers)
                yield chunk

            workbook.close()
        elif self.source_type in ['postgresql', 'mysql']:
            try:
                engine = create_db_engine(self.source_type, self.db_params)
                with engine.connect() as conn:
                    result = conn.execute(text(f"SELECT COUNT(*) FROM {self.db_params['table']}"))
                    self.total_rows_expected = result.fetchone()[0]
                self.logger.info(f"Total rows expected: {self.total_rows_expected}")

                conn = psycopg2.connect(**self.db_params) if self.source_type == 'postgresql' else mysql.connector.connect(**self.db_params)
                cursor = conn.cursor(name='etl_cursor' if self.source_type == 'postgresql' else None, buffered=False)
                cursor.execute(f"SELECT * FROM {self.db_params['table']}")
                headers = [desc[0] for desc in cursor.description]
                chunk_data = []

                while True:
                    rows = cursor.fetchmany(self.chunk_size)
                    if not rows:
                        break
                    chunk_data.extend(rows)
                    if len(chunk_data) >= self.chunk_size:
                        chunk = pd.DataFrame(chunk_data, columns=headers)
                        yield chunk
                        chunk_data = []

                if chunk_data:
                    chunk = pd.DataFrame(chunk_data, columns=headers)
                    yield chunk

                cursor.close()
                conn.close()
            except Exception as e:
                raise Exception(f"Database connection error: {e}")
        else:
            raise ValueError(f"Unsupported source type: {self.source_type}")

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
    def _migrate_chunk_to_db(self, chunk):
        try:
            if self.target_type == 'postgresql':
                conn = psycopg2.connect(
                    host=self.db_params_target['host'],
                    port=self.db_params_target['port'],
                    database=self.db_params_target['database'],
                    user=self.db_params_target['user'],
                    password=self.db_params_target['password']
                )
                cursor = conn.cursor()
                temp_csv = f"temp_chunk_{os.getpid()}.csv"
                chunk.to_csv(temp_csv, index=False, header=False, na_rep='\\N')

                columns = ",".join(chunk.columns)
                with open(temp_csv, 'r') as f:
                    cursor.copy_from(f, self.db_params_target['table'], sep=',', columns=chunk.columns, null='\\N')
                conn.commit()

                os.remove(temp_csv)
                cursor.close()
                conn.close()
            elif self.target_type == 'mysql':
                conn = mysql.connector.connect(
                    host=self.db_params_target['host'],
                    port=self.db_params_target['port'],
                    database=self.db_params_target['database'],
                    user=self.db_params_target['user'],
                    password=self.db_params_target['password']
                )
                cursor = conn.cursor()

                cursor.execute("SHOW VARIABLES LIKE 'local_infile'")
                local_infile = cursor.fetchone()
                if local_infile[1] != 'ON':
                    engine = create_db_engine(self.target_type, self.db_params_target)
                    chunk.to_sql(
                        self.db_params_target['table'],
                        engine,
                        if_exists='append',
                        index=False,
                        method='multi'
                    )
                else:
                    temp_csv = f"temp_chunk_{os.getpid()}.csv"
                    chunk.to_csv(temp_csv, index=False, header=False, na_rep='\\N')

                    columns = ",".join(chunk.columns)
                    cursor.execute(f"""
                        LOAD DATA LOCAL INFILE '{temp_csv}'
                        INTO TABLE {self.db_params_target['table']}
                        FIELDS TERMINATED BY ','
                        ENCLOSED BY '"'
                        LINES TERMINATED BY '\n'
                        ({columns})
                    """)
                    conn.commit()
                    os.remove(temp_csv)

                cursor.close()
                conn.close()
            else:
                raise ValueError(f"Unsupported target type: {self.target_type}")
        except Exception as e:
            self.logger.error(f"Ошибка миграции чанка в базу данных: {e}\n{traceback.format_exc()}")
            raise

    def run(self):
        self._predict_parameters_and_duration()
        self.logger.info(f"🚀 Starting ETL job: {self.source_type} -> {self.target_type}")
        start = time.time()
        chunk_num = 0

        try:
            if self.target_type == 'excel' and os.path.exists(self.target_file):
                workbook = load_workbook(self.target_file)
                worksheet = workbook.active
                headers_exist = any(worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1))
                if headers_exist or worksheet.max_row > 1:
                    workbook.close()
                    response = messagebox.askyesnocancel(
                        "Предупреждение",
                        f"Целевой файл {self.target_file} уже содержит данные. Перезаписать? (Да - перезаписать, Нет - добавить, Отмена - отменить)"
                    )
                    if response is None:
                        self.logger.warning("Migration cancelled by user.")
                        return
                    elif not response:
                        row = worksheet.max_row + 1 if worksheet.max_row is not None else 2
                    else:
                        row = 1
                else:
                    row = 1
                workbook.close()
            else:
                row = 1

            if self.target_type == 'excel':
                workbook = xlsxwriter.Workbook(self.target_file)
                worksheet = workbook.add_worksheet()

                if self.source_type == 'excel':
                    if not os.path.exists(self.source_file):
                        raise FileNotFoundError(f"Source file {self.source_file} does not exist.")
                    if not os.access(self.source_file, os.R_OK):
                        raise PermissionError(f"No read permission for {self.source_file}")
                    workbook_source = load_workbook(self.source_file, read_only=True)
                    sheet_source = workbook_source.active
                    headers = [cell.value for cell in next(sheet_source.rows)]
                    if not headers:
                        raise ValueError("Excel file is empty or has no headers")
                    self.total_rows_expected = sheet_source.max_row - 1
                    self.logger.info(f"Total rows expected: {self.total_rows_expected}")
                else:
                    stream = self.stream_data()
                    first_chunk = next(stream, None)
                    if first_chunk is None:
                        self.logger.warning("No data to migrate.")
                        workbook.close()
                        return
                    headers = list(first_chunk.columns)
                    stream = itertools.chain([first_chunk], stream)

                if row == 1:
                    for col_num, col_name in enumerate(headers):
                        worksheet.write(0, col_num, str(col_name) if col_name is not None else "")
                    row = 1
                else:
                    row = max(row - 1, 1)

                if self.source_type == 'excel':
                    chunk_data = []
                    for src_row in sheet_source.rows:
                        if src_row[0].row == 1:
                            continue
                        row_data = [cell.value for cell in src_row]
                        chunk_data.append(row_data)

                        if len(chunk_data) >= self.chunk_size:
                            for r, data_row in enumerate(chunk_data, start=row):
                                safe_row = [str(val) if val is not None else "" for val in data_row]
                                worksheet.write_row(r, 0, safe_row)
                            row += len(chunk_data)
                            self.total_rows += len(chunk_data)

                            if self.progress_callback and self.total_rows_expected > 0:
                                progress = (self.total_rows / self.total_rows_expected) * 100
                                self.progress_callback(progress)

                            chunk_num += 1
                            chunk_data = []

                    if chunk_data:
                        for r, data_row in enumerate(chunk_data, start=row):
                            safe_row = [str(val) if val is not None else "" for val in data_row]
                            worksheet.write_row(r, 0, safe_row)
                        self.total_rows += len(chunk_data)

                        if self.progress_callback and self.total_rows_expected > 0:
                            progress = (self.total_rows / self.total_rows_expected) * 100
                            self.progress_callback(progress)

                        chunk_num += 1

                    workbook_source.close()
                else:
                    for chunk in stream:
                        chunk_num += 1
                        for r, row_data in enumerate(chunk.values, start=row):
                            safe_row = [str(val) if pd.notna(val) else "" for val in row_data]
                            worksheet.write_row(r, 0, safe_row)
                        row += len(chunk)
                        self.total_rows += len(chunk)

                        if self.progress_callback and self.total_rows_expected > 0:
                            progress = (self.total_rows / self.total_rows_expected) * 100
                            self.progress_callback(progress)

                workbook.close()
                self.logger.info(f"📝 Migrated {self.total_rows} rows to {self.target_file}")
            else:
                stream = self.stream_data()
                first_chunk = next(stream, None)
                if first_chunk is None:
                    self.logger.warning("No data to migrate.")
                    return

                engine = create_db_engine(self.target_type, self.db_params_target)
                with engine.connect() as conn:
                    conn.execute(text(f"DROP TABLE IF EXISTS {self.db_params_target['table']}"))
                self.logger.info(f"🧹 Target table '{self.db_params_target['table']}' cleared.")
                first_chunk.to_sql(self.db_params_target['table'], engine, if_exists='replace', index=False)

                self.total_rows += len(first_chunk)
                if self.progress_callback and self.total_rows_expected > 0:
                    progress = (self.total_rows / self.total_rows_expected) * 100
                    self.progress_callback(progress)

                chunk_num += 1

                for chunk in stream:
                    chunk_num += 1
                    self._migrate_chunk_to_db(chunk)
                    self.total_rows += len(chunk)
                    if self.progress_callback and self.total_rows_expected > 0:
                        progress = (self.total_rows / self.total_rows_expected) * 100
                        self.progress_callback(progress)

            self.duration = round(time.time() - start, 2)
            self.logger.info(f"✅ ETL finished in {self.duration}s, {self.total_rows} rows migrated.")
            
            self._log_migration_data()
            self._train_model()
            validate_migration(self.config, self.logger)
        except Exception as e:
            self.logger.error(f"❌ ETL failed: {e}\n{traceback.format_exc()}")
            raise

# Загрузка конфигурации
def load_config(path='config.yaml'):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
    except FileNotFoundError:
        print(f"❌ Файл конфигурации {path} не найден.")
        sys.exit(1)
    except UnicodeDecodeError:
        print(f"❌ Файл конфигурации {path} имеет некорректную кодировку. Попробуйте сохранить его в UTF-8.")
        sys.exit(1)
    
    for key in config:
        if isinstance(config[key], str):
            config[key] = config[key].format(**os.environ)
    return config

# Основная функция миграции
def run_etl(config):
    job = ETLJob(config)
    job.run()
    return {"status": "success", "rows_migrated": job.total_rows, "duration": job.duration}

# Диалог для ввода параметров базы данных
class DatabaseDialog(tk.Toplevel):
    def __init__(self, parent, title, db_type):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.db_type = db_type
        self.geometry("400x400")

        tk.Label(self, text="Host:").pack()
        self.host_entry = tk.Entry(self)
        self.host_entry.insert(0, "localhost")
        self.host_entry.pack()

        tk.Label(self, text="Port:").pack()
        self.port_entry = tk.Entry(self)
        self.port_entry.insert(0, "5432" if "postgresql" in db_type.lower() else "3306")
        self.port_entry.pack()

        tk.Label(self, text="Database:").pack()
        self.db_entry = tk.Entry(self)
        self.db_entry.pack()

        tk.Label(self, text="User:").pack()
        self.user_entry = tk.Entry(self)
        self.user_entry.pack()

        tk.Label(self, text="Password:").pack()
        self.password_entry = tk.Entry(self, show="*")
        self.password_entry.pack()

        tk.Label(self, text="Table:").pack()
        self.table_entry = tk.Entry(self)
        self.table_entry.pack()

        tk.Button(self, text="OK", command=self.on_ok).pack(pady=10)
        tk.Button(self, text="Cancel", command=self.on_cancel).pack()

    def on_ok(self):
        self.result = {
            "host": self.host_entry.get(),
            "port": self.port_entry.get(),
            "database": self.db_entry.get(),
            "user": self.user_entry.get(),
            "password": self.password_entry.get(),
            "table": self.table_entry.get()
        }
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()

# Графический интерфейс
class ETLApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ETL Migration Tool")
        self.logger = get_logger()
        self.config = load_config()
        self.source_db_params = None
        self.target_db_params = None

        tk.Label(root, text="ETL Migration Tool", font=("Arial", 16)).pack(pady=10)

        tk.Label(root, text="Тип источника:").pack()
        self.source_type = tk.StringVar(value="excel")
        source_types = ["excel", "postgresql", "mysql"]
        self.source_type_menu = ttk.Combobox(root, textvariable=self.source_type, values=source_types, state="readonly")
        self.source_type_menu.pack()
        self.source_type_menu.bind("<<ComboboxSelected>>", self.on_source_type_change)

        tk.Label(root, text="Исходный файл или база данных:").pack()
        self.source_file_entry = tk.Entry(root, width=50)
        self.source_file_entry.pack()
        self.source_file_button = tk.Button(root, text="Выбрать исходный файл", command=self.select_source_file)
        self.source_file_button.pack(pady=5)

        tk.Label(root, text="Тип цели:").pack()
        self.target_type = tk.StringVar(value="excel")
        target_types = ["excel", "postgresql", "mysql"]
        self.target_type_menu = ttk.Combobox(root, textvariable=self.target_type, values=target_types, state="readonly")
        self.target_type_menu.pack()
        self.target_type_menu.bind("<<ComboboxSelected>>", self.on_target_type_change)

        tk.Label(root, text="Целевой файл или база данных:").pack()
        self.target_file_entry = tk.Entry(root, width=50)
        self.target_file_entry.pack()
        self.target_file_button = tk.Button(root, text="Выбрать целевой файл", command=self.select_target_file)
        self.target_file_button.pack(pady=5)

        tk.Button(root, text="Настроить подключение к БД", command=self.open_db_config_dialog).pack(pady=5)

        tk.Label(root, text="Chunk Size:").pack()
        self.chunk_size_entry = tk.Entry(root)
        self.chunk_size_entry.insert(0, str(self.config.get("chunk_size", 10000)))
        self.chunk_size_entry.pack()

        tk.Label(root, text="Commit Interval:").pack()
        self.commit_interval_entry = tk.Entry(root)
        self.commit_interval_entry.insert(0, str(self.config.get("commit_interval", 20000)))
        self.commit_interval_entry.pack()

        self.predicted_time_label = tk.Label(root, text="Предполагаемое время: -")
        self.predicted_time_label.pack(pady=5)

        tk.Button(root, text="Запустить миграцию", command=self.run_migration).pack(pady=5)
        tk.Button(root, text="Обучить модель", command=self.train_model).pack(pady=5)

        self.progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(pady=10)

        self.log_area = scrolledtext.ScrolledText(root, width=80, height=20)
        self.log_area.pack(pady=10)

    def log(self, message):
        def update_log():
            self.log_area.insert(tk.END, message + "\n")
            self.log_area.yview(tk.END)
        self.root.after(0, update_log)
        self.logger.info(message)

    def on_source_type_change(self, event):
        source_type = self.source_type.get()
        if source_type == 'excel':
            self.source_file_button.config(text="Выбрать исходный файл", command=self.select_source_file)
        else:
            self.source_file_button.config(text="Выбрать базу данных", command=self.select_source_db)

    def on_target_type_change(self, event):
        target_type = self.target_type.get()
        if target_type == 'excel':
            self.target_file_button.config(text="Выбрать целевой файл", command=self.select_target_file)
        else:
            self.target_file_button.config(text="Выбрать базу данных", command=self.select_target_db)

    def select_source_file(self):
        file_path = filedialog.askopenfilename(
            title="Выберите исходный файл",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            if not os.path.exists(file_path):
                messagebox.showerror("Ошибка", f"Файл {file_path} не существует.")
                return
            if not os.access(file_path, os.R_OK):
                messagebox.showerror("Ошибка", f"Нет прав на чтение файла {file_path}.")
                return
            self.source_file_entry.delete(0, tk.END)
            self.source_file_entry.insert(0, file_path)
            self.log(f"Выбран исходный файл: {file_path}")

    def select_target_file(self):
        file_path = filedialog.asksaveasfilename(
            title="Выберите целевой файл",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            defaultextension=".xlsx"
        )
        if file_path:
            if os.path.exists(file_path):
                if not os.access(file_path, os.W_OK):
                    messagebox.showerror("Ошибка", f"Нет прав на запись в файл {file_path}.")
                    return
            self.target_file_entry.delete(0, tk.END)
            self.target_file_entry.insert(0, file_path)
            self.log(f"Выбран целевой файл: {file_path}")

    def select_source_db(self):
        dialog = DatabaseDialog(self.root, f"Параметры {self.source_type.get().upper()} источника", self.source_type.get())
        self.root.wait_window(dialog)
        if dialog.result:
            self.source_file_entry.delete(0, tk.END)
            self.source_file_entry.insert(0, f"{self.source_type.get()}://{dialog.result['host']}:{dialog.result['port']}/{dialog.result['database']}")
            self.source_db_params = dialog.result
            self.log(f"Выбрана база данных источника: {self.source_file_entry.get()}")

    def select_target_db(self):
        dialog = DatabaseDialog(self.root, f"Параметры {self.target_type.get().upper()} цели", self.target_type.get())
        self.root.wait_window(dialog)
        if dialog.result:
            self.target_file_entry.delete(0, tk.END)
            self.target_file_entry.insert(0, f"{self.target_type.get()}://{dialog.result['host']}:{dialog.result['port']}/{dialog.result['database']}")
            self.target_db_params = dialog.result
            self.log(f"Выбрана база данных цели: {self.target_file_entry.get()}")

    def open_db_config_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Настройка подключения к БД")
        dialog.geometry("600x400")

        notebook = ttk.Notebook(dialog)
        notebook.pack(fill="both", expand=True)

        source_frame = ttk.Frame(notebook)
        notebook.add(source_frame, text="Источник")
        tk.Label(source_frame, text=f"Параметры источника ({self.source_type.get().upper()}):").pack(pady=5)
        source_dialog = DatabaseDialog(source_frame, "Источник", self.source_type.get())
        source_dialog.pack(fill="both", expand=True)

        target_frame = ttk.Frame(notebook)
        notebook.add(target_frame, text="Цель")
        tk.Label(target_frame, text=f"Параметры цели ({self.target_type.get().upper()}):").pack(pady=5)
        target_dialog = DatabaseDialog(target_frame, "Цель", self.target_type.get())
        target_dialog.pack(fill="both", expand=True)

        def save_db_config():
            self.source_db_params = source_dialog.result
            self.target_db_params = target_dialog.result
            if self.source_db_params:
                self.source_file_entry.delete(0, tk.END)
                self.source_file_entry.insert(0, f"{self.source_type.get()}://{self.source_db_params['host']}:{self.source_db_params['port']}/{self.source_db_params['database']}")
                self.log(f"Настроены параметры источника: {self.source_file_entry.get()}")
            if self.target_db_params:
                self.target_file_entry.delete(0, tk.END)
                self.target_file_entry.insert(0, f"{self.target_type.get()}://{self.target_db_params['host']}:{self.target_db_params['port']}/{self.target_db_params['database']}")
                self.log(f"Настроены параметры цели: {self.target_file_entry.get()}")
            dialog.destroy()

        tk.Button(dialog, text="Сохранить", command=save_db_config).pack(pady=10)
        tk.Button(dialog, text="Отмена", command=dialog.destroy).pack()

    def run_migration(self):
        threading.Thread(target=self._run_migration, daemon=True).start()

    def train_model(self):
        threading.Thread(target=self._train_model_thread, daemon=True).start()

    def _train_model_thread(self):
        self.log("Запуск обучения модели...")
        try:
            job = ETLJob(self.config)
            job._train_model()
            self.log("Обучение модели завершено.")
        except Exception as e:
            self.log(f"❌ Ошибка обучения модели: {str(e)}\n{traceback.format_exc()}")
            messagebox.showerror("Ошибка", f"Ошибка обучения модели: {str(e)}")

    def _run_migration(self):
        source_type = self.source_type.get()
        target_type = self.target_type.get()
        source_file = self.source_file_entry.get()
        target_file = self.target_file_entry.get()
        if not source_file or not target_file:
            messagebox.showwarning("Предупреждение", "Выберите исходный и целевой файлы или базы данных.")
            return
        self.log("Запуск миграции...")
        self.progress["value"] = 0
        try:
            config = self.config.copy()
            config['source_type'] = source_type
            config['target_type'] = target_type
            config['source_file'] = source_file
            config['target_file'] = target_file
            config['db_params'] = self.source_db_params if self.source_db_params else config.get('db_params', {})
            config['db_params_target'] = self.target_db_params if self.target_db_params else config.get('db_params_target', {})
            config['chunk_size'] = int(self.chunk_size_entry.get())
            config['commit_interval'] = int(self.commit_interval_entry.get())
            config['ga_pop_size'] = self.config.get('ga_pop_size', 20)
            config['ga_generations'] = self.config.get('ga_generations', 10)
            config['ga_weights'] = self.config.get('ga_weights', [10, 2, 2, 2, 0.5, 0.5])

            job = ETLJob(config)

            def update_progress(value):
                def update():
                    self.progress["value"] = value
                    if value > 0 and job.predicted_duration > 0:
                        remaining = (100 - value) * job.predicted_duration / 100
                        self.predicted_time_label.config(text=f"Предполагаемое время: {int(remaining)}s осталось")
                self.root.after(0, update)

            config['progress_callback'] = update_progress
            self.predicted_time_label.config(text=f"Предполагаемое время: {job.predicted_duration}s")
            result = run_etl(config)

            self.log(f"Миграция завершена: {result}")
            self.progress["value"] = 100
            self.predicted_time_label.config(text="Предполагаемое время: Завершено")
        except Exception as e:
            self.log(f"❌ Миграция не удалась: {str(e)}\n{traceback.format_exc()}")
            messagebox.showerror("Ошибка", f"Миграция не удалась: {str(e)}")

# Точка входа
if __name__ == "__main__":
    root = tk.Tk()
    app = ETLApp(root)
    root.mainloop()